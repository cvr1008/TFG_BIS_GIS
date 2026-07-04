import json
import tempfile
import unittest
from datetime import datetime, timedelta
from pathlib import Path
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook, load_workbook

from src.pacientes_icca import (
    generar_icca_sintetico,
    listar_pacientes,
    preparar_auditoria_gasometrias,
    preparar_sesiones_paciente,
)
from src.visualizacion_icca import cargar_datos_icca


class PacientesIccaTest(unittest.TestCase):
    def _crear_icca(self, ruta, inicio):
        libro = Workbook()
        hoja = libro.active
        hoja.title = "constantes_vitales"
        hoja.append(["Hoja 2"])
        hoja.append([])
        hoja.append(
            [
                "paciente_id",
                "sesion_bis_id",
                "timestamp",
                "fuente_pdf",
                "bloque_origen",
                "variable",
                "valor",
                "unidad",
                "pa_sistolica_mmHg",
                "pa_diastolica_mmHg",
                "pa_media_mmHg",
                "detalle_original",
                "origen_pdf",
                "observaciones",
            ]
        )
        for segundos, fc, sistolica, diastolica, media in (
            (5, 70, 110, 60, 77),
            (10, 80, 120, 70, 87),
        ):
            instante = inicio + timedelta(seconds=segundos)
            hoja.append(
                ["PACIENTE_001", "BIS01", instante, "Ventilatoria", "Constantes", "FC", fc, "lpm"]
            )
            hoja.append(
                [
                    "PACIENTE_001",
                    "BIS01",
                    instante,
                    "Ventilatoria",
                    "Constantes",
                    "Presion arterial",
                    None,
                    "mmHg",
                    sistolica,
                    diastolica,
                    media,
                ]
            )
        analisis = libro.create_sheet("analisis")
        analisis.append(["Hoja 3"])
        analisis.append([])
        analisis.append(["timestamp", "variable", "valor", "unidad"])
        libro.save(ruta)

    def test_sintesis_conserva_reales_y_no_extrapola(self):
        with tempfile.TemporaryDirectory() as temporal:
            raiz = Path(temporal)
            origen = raiz / "ICCA.xlsx"
            salida = raiz / "ICCA_sintetico.csv"
            inicio = datetime(2026, 5, 1, 10, 0, 0)
            self._crear_icca(origen, inicio)

            resumen = generar_icca_sintetico(
                origen,
                inicio,
                inicio + timedelta(seconds=15),
                "PACIENTE_001",
                "BIS01",
                salida,
            )
            datos = pd.read_csv(salida)
            columna = "fc__valor"

            self.assertEqual(resumen["filas"], 16)
            self.assertTrue(pd.isna(datos.loc[0, columna]))
            self.assertEqual(datos.loc[5, columna], 70)
            self.assertIn("fc", datos.loc[5, "series_reales"])
            self.assertEqual(datos.loc[10, columna], 80)
            self.assertIn("fc", datos.loc[10, "series_reales"])
            self.assertTrue(pd.isna(datos.loc[7, "series_reales"]))
            self.assertTrue(pd.isna(datos.loc[15, columna]))
            metadata = json.loads(
                salida.with_suffix(".meta.json").read_text(encoding="utf-8")
            )
            self.assertEqual(metadata["gasometrias_auditoria"], [])

    def test_sintesis_es_reproducible(self):
        with tempfile.TemporaryDirectory() as temporal:
            raiz = Path(temporal)
            origen = raiz / "ICCA.xlsx"
            inicio = datetime(2026, 5, 1, 10, 0, 0)
            self._crear_icca(origen, inicio)
            salidas = [raiz / "a.csv", raiz / "b.csv"]
            for salida in salidas:
                generar_icca_sintetico(
                    origen,
                    inicio,
                    inicio + timedelta(seconds=15),
                    "PACIENTE_001",
                    "BIS01",
                    salida,
                )
            a = pd.read_csv(salidas[0])
            b = pd.read_csv(salidas[1])
            pd.testing.assert_frame_equal(a, b)

    def test_lista_carpetas_con_manifiesto(self):
        with tempfile.TemporaryDirectory() as temporal:
            carpeta = Path(temporal) / "PACIENTE_001"
            carpeta.mkdir()
            (carpeta / "paciente.json").write_text(
                json.dumps({"paciente_id": "PACIENTE_001", "sesiones": []}),
                encoding="utf-8",
            )
            pacientes = listar_pacientes(temporal)
            self.assertEqual([item["paciente_id"] for item in pacientes], ["PACIENTE_001"])

    def test_prepara_sesion_bis_sin_icca(self):
        with tempfile.TemporaryDirectory() as temporal:
            raiz = Path(temporal)
            paciente = raiz / "PACIENTE_001"
            carpeta_bis = paciente / "SESIONES" / "SESION_001_BIS01" / "BIS" / "BIS01"
            carpeta_bis.mkdir(parents=True)
            manifiesto = {
                "paciente_id": "PACIENTE_001",
                "sesiones": [
                    {
                        "nombre_carpeta": "SESION_001_BIS01",
                        "sesion_bis_id": "BIS01",
                        "inicio_bis": "2026-05-01 10:00:00",
                        "fin_bis": "2026-05-01 11:00:00",
                        "carpeta_bis": str(carpeta_bis.relative_to(paciente)),
                        "excel_icca_auxiliar": None,
                        "solapamientos": [],
                    }
                ],
            }
            (paciente / "paciente.json").write_text(
                json.dumps(manifiesto), encoding="utf-8"
            )

            with patch(
                "src.pacientes_icca._detectar_bis_con_cache",
                return_value={
                    "validacion": {},
                    "fa_disponible": True,
                    "origenes": ["fa"],
                },
            ):
                sesiones = preparar_sesiones_paciente(
                    raiz, "PACIENTE_001", generar_sinteticos=True
                )

            self.assertEqual(len(sesiones), 1)
            self.assertEqual(sesiones[0]["estado_icca"], "ausente")
            self.assertFalse(sesiones[0]["icca_disponible"])
            self.assertIsNone(sesiones[0]["icca_auxiliar_absoluto"])
            self.assertNotIn("icca_sintetico", sesiones[0])

    def test_gasometria_explicita_admite_medicion_aislada(self):
        datos = pd.DataFrame(
            [
                {
                    "timestamp": datetime(2026, 5, 1, 10, 0),
                    "fuente_pdf": "Ventilatoria",
                    "variable": "pO2",
                    "valor": 92,
                    "tipo_gasometria": "Arterial",
                }
            ]
        )
        auditoria = preparar_auditoria_gasometrias(datos)
        self.assertEqual(auditoria.loc[0, "incluida_visualizacion"], "si")
        self.assertEqual(auditoria.loc[0, "tipo_gasometria_final"], "arterial")
        self.assertEqual(auditoria.loc[0, "origen_tipo_gasometria"], "indicado")

    def test_gasometria_no_tipificada_aislada_se_excluye(self):
        datos = pd.DataFrame(
            [
                {
                    "timestamp": datetime(2026, 5, 1, 10, 0),
                    "fuente_pdf": "Ventilatoria",
                    "variable": "pO2",
                    "valor": 45,
                }
            ]
        )
        auditoria = preparar_auditoria_gasometrias(datos)
        self.assertEqual(auditoria.loc[0, "incluida_visualizacion"], "no")
        self.assertIn("gasometria incompleta", auditoria.loc[0, "motivo_exclusion"])

    def test_gasometria_no_tipificada_completa_infiere_venosa(self):
        instante = datetime(2026, 5, 1, 10, 0)
        datos = pd.DataFrame(
            [
                {
                    "timestamp": instante,
                    "fuente_pdf": "Ventilatoria",
                    "bloque_origen": "Gasometria",
                    "variable": "pO2",
                    "valor": 45,
                },
                {
                    "timestamp": instante,
                    "fuente_pdf": "Ventilatoria",
                    "bloque_origen": "Gasometria",
                    "variable": "pCO2",
                    "valor": 38,
                    "marca_original": "^",
                },
            ]
        )
        auditoria = preparar_auditoria_gasometrias(datos)
        self.assertEqual(set(auditoria["incluida_visualizacion"]), {"si"})
        self.assertEqual(set(auditoria["tipo_gasometria_final"]), {"venosa"})
        self.assertEqual(set(auditoria["confianza_inferencia"]), {"alta"})

    def test_gasometria_con_indicios_contradictorios_se_excluye(self):
        instante = datetime(2026, 5, 1, 10, 0)
        datos = pd.DataFrame(
            [
                {
                    "timestamp": instante,
                    "fuente_pdf": "Ventilatoria",
                    "variable": "pO2",
                    "valor": 45,
                },
                {
                    "timestamp": instante,
                    "fuente_pdf": "Ventilatoria",
                    "variable": "pCO2",
                    "valor": 50,
                    "marca_original": "^",
                },
            ]
        )
        auditoria = preparar_auditoria_gasometrias(datos)
        self.assertEqual(set(auditoria["incluida_visualizacion"]), {"no"})
        self.assertTrue(
            auditoria["motivo_exclusion"].str.contains("contradictorios").all()
        )

    def test_visualizador_oculta_gasometria_excluida_y_conserva_otros_analisis(self):
        with tempfile.TemporaryDirectory() as temporal:
            raiz = Path(temporal)
            origen = raiz / "ICCA.xlsx"
            salida = raiz / "ICCA_sintetico.csv"
            inicio = datetime(2026, 5, 1, 10, 0, 0)
            self._crear_icca(origen, inicio)
            libro = load_workbook(origen)
            hoja = libro["analisis"]
            hoja.append([inicio + timedelta(seconds=4), "Sodio", 140, "mEq/L"])
            hoja.append([inicio + timedelta(seconds=5), "pO2", 45, "mmHg"])
            libro.save(origen)

            generar_icca_sintetico(
                origen,
                inicio,
                inicio + timedelta(seconds=15),
                "PACIENTE_001",
                "BIS01",
                salida,
            )
            datos = cargar_datos_icca(salida, origen)
            variables = set(datos["analisis"]["variable"].map(str))
            self.assertIn("Sodio", variables)
            self.assertNotIn("pO2", variables)


if __name__ == "__main__":
    unittest.main()
