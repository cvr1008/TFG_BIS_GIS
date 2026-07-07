import json
import gc
import tempfile
import unittest
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table

from src.intervalos import (
    analizar_seleccion,
    descubrir_icca_en_carpeta,
    descubrir_sesiones_bis_en_carpeta,
    leer_intervalo_bis,
    leer_intervalo_icca,
)
from src.pacientes import (
    actualizar_paciente,
    crear_paciente,
    eliminar_paciente,
    listar_pacientes,
    obtener_fuentes_paciente,
)


def crear_spa(carpeta, inicio, segundos=4):
    carpeta.mkdir(parents=True, exist_ok=True)
    spa = carpeta / "L04301923.spa"
    lineas = ["Cabecera BIS", "Time|BIS"]
    for desplazamiento in range(segundos):
        instante = inicio.replace(microsecond=0)
        instante = instante.fromtimestamp(instante.timestamp() + desplazamiento)
        lineas.append(f"{instante.strftime('%m/%d/%Y %H:%M:%S')}|50")
    spa.write_text("\n".join(lineas), encoding="latin1")
    (carpeta / "L04301923.r4a").write_bytes(b"raw")
    return spa


def crear_icca(ruta, ingreso, alta):
    libro = Workbook()
    general = libro.active
    general.title = "general"
    general.append(["Hoja general"])
    general.append([])
    general.append([
        "paciente_id",
        "carpeta_paciente",
        "sesion_bis_id",
        "fecha_hora_ingreso",
        "fecha_hora_alta",
    ])
    general.append(["", "", "", ingreso, alta])

    for nombre in ("constantes_vitales", "analisis", "perfusiones"):
        hoja = libro.create_sheet(nombre)
        hoja.append([nombre])
        hoja.append([])
        hoja.append(["paciente_id", "sesion_bis_id", "timestamp", "variable", "valor"])
        hoja.append(["", "", ingreso, "antes", 1])
        hoja.append(["", "", datetime(2026, 4, 30, 19, 23, 30), "dentro", 2])
        hoja.append(["", "", alta, "despues", 3])
        tabla = Table(displayName=f"Tabla_{nombre}", ref="A3:E6")
        hoja.add_table(tabla)

    libro.save(ruta)
    libro.close()


def crear_icca_perfusiones(ruta, ingreso, alta):
    libro = Workbook()
    general = libro.active
    general.title = "general"
    general.append(["Hoja general"])
    general.append([])
    general.append([
        "paciente_id",
        "carpeta_paciente",
        "sesion_bis_id",
        "fecha_hora_ingreso",
        "fecha_hora_alta",
    ])
    general.append(["", "", "", ingreso, alta])

    for nombre in ("constantes_vitales", "analisis"):
        hoja = libro.create_sheet(nombre)
        hoja.append([nombre])
        hoja.append([])
        hoja.append(["paciente_id", "sesion_bis_id", "timestamp", "variable", "valor"])
        hoja.append(["", "", ingreso, "dentro", 2])
        hoja.add_table(Table(displayName=f"Tabla_{nombre}", ref="A3:E4"))

    perfusiones = libro.create_sheet("perfusiones")
    perfusiones.append(["perfusiones"])
    perfusiones.append([])
    perfusiones.append([
        "paciente_id",
        "sesion_bis_id",
        "timestamp",
        "farmaco",
        "dosis_actual",
        "unidad_dosis",
        "velocidad_bomba_ml_h",
        "volumen_acumulado_24h_ml",
        "detalle_original",
    ])
    perfusiones.append(["", "", datetime(2026, 4, 30, 7, 0, 0), "Propofol", 1, "mg/h", 10, None, "1 >> 10"])
    perfusiones.append(["", "", datetime(2026, 4, 30, 9, 0, 0), "Propofol", 1, "mg/h", 10, None, "1 >> 10"])
    perfusiones.append(["", "", datetime(2026, 4, 30, 10, 0, 0), "Propofol", 1, "mg/h", 10, 30, "1 >> 10 (30)"])
    perfusiones.add_table(Table(displayName="Tabla_perfusiones", ref="A3:I6"))
    libro.save(ruta)
    libro.close()


class FlujoOrganizadorTest(unittest.TestCase):
    def setUp(self):
        self.temporal = tempfile.TemporaryDirectory()
        self.raiz = Path(self.temporal.name)
        self.bis = self.raiz / "BIS_ORIGEN"
        self.inicio_bis = datetime(2026, 4, 30, 19, 23, 29)
        crear_spa(self.bis, self.inicio_bis, segundos=4)
        self.icca = self.raiz / "icca.xlsx"
        crear_icca(
            self.icca,
            datetime(2026, 4, 29, 13, 0, 0),
            datetime(2026, 5, 20, 17, 46, 0),
        )

    def tearDown(self):
        gc.collect()
        self.temporal.cleanup()

    def test_intervalos_y_solapamiento(self):
        intervalo_bis = leer_intervalo_bis(self.bis)
        intervalo_icca = leer_intervalo_icca(self.icca)
        self.assertEqual(intervalo_bis["inicio"], "2026-04-30 19:23:29")
        self.assertEqual(intervalo_bis["fin"], "2026-04-30 19:23:32")
        self.assertEqual(intervalo_icca["inicio"], "2026-04-29 13:00:00")

        analisis = analizar_seleccion([self.icca], [self.bis])
        self.assertTrue(analisis["bis"][0]["icca_compatibles"])
        self.assertTrue(analisis["bis"][0]["solapamientos"][0]["completo"])

    def test_descubre_carpetas_madre(self):
        madre_bis = self.raiz / "MADRE_BIS"
        crear_spa(madre_bis / "sesion_a", datetime(2026, 4, 30, 19, 0, 0), segundos=3)
        crear_spa(madre_bis / "sesion_b", datetime(2026, 4, 30, 20, 0, 0), segundos=3)
        madre_icca = self.raiz / "MADRE_ICCA"
        madre_icca.mkdir()
        icca_copia = madre_icca / "icca_copia.xlsx"
        crear_icca(icca_copia, datetime(2026, 4, 29, 13, 0, 0), datetime(2026, 5, 1, 0, 0, 0))

        sesiones = descubrir_sesiones_bis_en_carpeta(madre_bis)
        iccas = descubrir_icca_en_carpeta(madre_icca)

        self.assertEqual(len(sesiones), 2)
        self.assertEqual(iccas, [str(icca_copia.resolve())])

    def test_no_permite_bis_solapados_en_mismo_paciente(self):
        salida = self.raiz / "PACIENTES"
        bis_largo = self.raiz / "BIS_LARGO"
        crear_spa(bis_largo, self.inicio_bis, segundos=60)
        bis_solapado = self.raiz / "BIS_SOLAPADO"
        crear_spa(
            bis_solapado,
            self.inicio_bis + timedelta(seconds=29),
            segundos=4,
        )

        with self.assertRaisesRegex(ValueError, "No se pueden asignar dos sesiones BIS solapadas"):
            crear_paciente(salida, [self.icca], [bis_largo, bis_solapado])

    def test_permite_solape_bis_menor_de_30_segundos(self):
        salida = self.raiz / "PACIENTES"
        bis_largo = self.raiz / "BIS_LARGO"
        crear_spa(bis_largo, self.inicio_bis, segundos=60)
        bis_solape_pequeno = self.raiz / "BIS_SOLAPE_PEQUENO"
        crear_spa(
            bis_solape_pequeno,
            self.inicio_bis + timedelta(seconds=30),
            segundos=4,
        )

        manifiesto = crear_paciente(
            salida,
            [self.icca],
            [bis_largo, bis_solape_pequeno],
        )

        self.assertEqual(len(manifiesto["sesiones"]), 2)

    def test_permite_bis_contiguos_con_mismo_segundo_fin_inicio(self):
        salida = self.raiz / "PACIENTES"
        bis_contiguo = self.raiz / "BIS_CONTIGUO"
        crear_spa(
            bis_contiguo,
            self.inicio_bis + timedelta(seconds=3),
            segundos=4,
        )

        manifiesto = crear_paciente(salida, [], [self.bis, bis_contiguo])

        self.assertEqual(len(manifiesto["sesiones"]), 2)
        self.assertEqual(
            manifiesto["sesiones"][0]["fin_bis"],
            manifiesto["sesiones"][1]["inicio_bis"],
        )

    def test_crea_paciente_y_recorta_excel(self):
        salida = self.raiz / "PACIENTES"
        manifiesto = crear_paciente(salida, [self.icca], [self.bis])

        self.assertEqual(manifiesto["paciente_id"], "PACIENTE_001")
        carpeta = salida / "PACIENTE_001"
        self.assertTrue((carpeta / "paciente.json").is_file())
        datos = json.loads((carpeta / "paciente.json").read_text(encoding="utf-8"))
        self.assertEqual(len(datos["sesiones"]), 1)

        auxiliar = carpeta / datos["sesiones"][0]["excel_icca_auxiliar"]
        self.assertTrue(auxiliar.is_file())
        libro = load_workbook(auxiliar, read_only=True, data_only=True)
        try:
            hoja = libro["analisis"]
            filas = list(hoja.iter_rows(min_row=4, values_only=True))
            filas_validas = [fila for fila in filas if fila[2] is not None]
            self.assertEqual(len(filas_validas), 1)
            self.assertEqual(filas_validas[0][3], "dentro")
            self.assertEqual(filas_validas[0][0], "PACIENTE_001")
            self.assertEqual(filas_validas[0][1], "L04301923")
        finally:
            libro.close()

    def test_excel_auxiliar_conserva_dosis_perfusiones_sin_acumulado_calculado(self):
        salida = self.raiz / "PACIENTES"
        icca = self.raiz / "icca_perfusiones.xlsx"
        crear_icca_perfusiones(
            icca,
            datetime(2026, 4, 30, 6, 0, 0),
            datetime(2026, 4, 30, 12, 0, 0),
        )
        bis = self.raiz / "BIS_PERFUSION"
        crear_spa(bis, datetime(2026, 4, 30, 6, 30, 0), segundos=14401)

        manifiesto = crear_paciente(salida, [icca], [bis])
        carpeta = salida / manifiesto["paciente_id"]
        auxiliar = carpeta / manifiesto["sesiones"][0]["excel_icca_auxiliar"]
        libro = load_workbook(auxiliar, read_only=True, data_only=True)
        try:
            hoja = libro["perfusiones"]
            cabeceras = [celda.value for celda in hoja[3]]
            self.assertNotIn("volumen_acumulado_calculado_ml", cabeceras)
            self.assertNotIn("dia_clinico_inicio", cabeceras)
            self.assertNotIn("acumulado_calculado_origen", cabeceras)
            indice_dosis = cabeceras.index("dosis_actual")
            filas = [
                fila
                for fila in hoja.iter_rows(min_row=4, values_only=True)
                if fila[2] is not None
            ]
            dosis = [fila[indice_dosis] for fila in filas]
            self.assertEqual(dosis, [1, 1, 1])
        finally:
            libro.close()

    def test_crea_paciente_con_sesion_bis_sin_icca(self):
        salida = self.raiz / "PACIENTES"
        manifiesto = crear_paciente(salida, [], [self.bis])

        self.assertEqual(manifiesto["icca"], [])
        self.assertEqual(manifiesto["fuentes"]["icca"], [])
        self.assertEqual(len(manifiesto["sesiones"]), 1)
        sesion = manifiesto["sesiones"][0]
        self.assertIsNone(sesion["excel_icca_auxiliar"])
        self.assertEqual(sesion["icca_origen"], [])
        self.assertEqual(sesion["solapamientos"], [])
        carpeta = salida / manifiesto["paciente_id"]
        self.assertTrue((carpeta / sesion["carpeta_bis"]).is_dir())
        self.assertFalse(
            (carpeta / "SESIONES" / sesion["nombre_carpeta"] / "ICCA").exists()
        )

    def test_no_permite_reutilizar_los_mismos_archivos(self):
        salida = self.raiz / "PACIENTES"
        crear_paciente(salida, [self.icca], [self.bis])

        with self.assertRaisesRegex(
            ValueError,
            r"Archivo ICCA y sesión BIS ya asignados a un paciente \(PACIENTE_001\)",
        ):
            crear_paciente(salida, [self.icca], [self.bis])

        self.assertEqual(len(listar_pacientes(salida)), 1)

    def test_actualiza_el_mismo_paciente_y_permite_eliminarlo(self):
        salida = self.raiz / "PACIENTES"
        crear_paciente(salida, [self.icca], [self.bis])
        icca_nuevo = self.raiz / "icca_corregido.xlsx"
        crear_icca(
            icca_nuevo,
            datetime(2026, 4, 29, 13, 0, 0),
            datetime(2026, 5, 20, 17, 46, 0),
        )

        actualizado = actualizar_paciente(
            salida,
            "PACIENTE_001",
            [icca_nuevo],
            [self.bis],
        )

        self.assertEqual(actualizado["paciente_id"], "PACIENTE_001")
        fuentes = obtener_fuentes_paciente(actualizado)
        self.assertEqual(fuentes["icca"], [str(icca_nuevo.resolve())])
        self.assertEqual(len(actualizado["sesiones"]), 1)
        auxiliar = salida / "PACIENTE_001" / actualizado["sesiones"][0]["excel_icca_auxiliar"]
        self.assertTrue(auxiliar.is_file())

        eliminar_paciente(salida, "PACIENTE_001")
        self.assertEqual(listar_pacientes(salida), [])


if __name__ == "__main__":
    unittest.main()
