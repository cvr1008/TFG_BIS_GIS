from __future__ import annotations

from contextlib import contextmanager
from pathlib import Path
from shutil import copy2
from tempfile import TemporaryDirectory

from openpyxl import load_workbook


@contextmanager
def abrir_libro(ruta, **opciones):
    """
    Abre un archivo Excel como un libro de openpyxl.

    La función intenta abrir directamente el archivo indicado. Si el archivo
    está bloqueado por otro programa, por ejemplo porque está abierto en Excel,
    se crea una copia temporal y se abre dicha copia en modo lectura. Esto
    permite consultar el contenido del libro sin modificar ni desbloquear el
    archivo original.

    Parámetros
    ----------
    ruta : str | pathlib.Path
        Ruta del archivo Excel que se desea abrir.

    **opciones
        Opciones adicionales que se pasan directamente a `openpyxl.load_workbook`.
        Por ejemplo, `read_only=True`, `data_only=True` o `keep_vba=True`.

    Devuelve
    --------
    openpyxl.workbook.workbook.Workbook
        Libro de Excel abierto mediante openpyxl. Se devuelve dentro del bloque
        `with`, ya que la función actúa como gestor de contexto.

    """
    ruta = Path(ruta).expanduser().resolve()
    temporal = None
    try:
        try:
            libro = load_workbook(ruta, **opciones)
        except PermissionError:
            temporal = TemporaryDirectory(prefix="icca_excel_")
            copia = Path(temporal.name) / ruta.name
            copy2(ruta, copia)
            libro = load_workbook(copia, **opciones)

        try:
            yield libro
        finally:
            libro.close()
    finally:
        if temporal is not None:
            temporal.cleanup()
