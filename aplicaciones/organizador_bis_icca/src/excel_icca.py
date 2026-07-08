from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from src.intervalos import _parsear_fecha
from src.libros import abrir_libro


HOJAS_TEMPORALES = ("constantes_vitales", "analisis", "perfusiones")
def _buscar_hoja(libro, nombre):
    """
    Busca hoja.

    Parámetros
    ----------
    libro : Any
        Valor de entrada utilizado por la función.

    nombre : Any
        Valor de entrada utilizado por la función.

    Devuelve
    --------
    Any
        Resultado generado por la función.
    """
    objetivo = nombre.casefold()
    for hoja in libro.worksheets:
        if hoja.title.casefold() == objetivo:
            return hoja
    return None


def _cabeceras(hoja):
    """
    Ejecuta la lógica asociada a cabeceras.

    Parámetros
    ----------
    hoja : Any
        Valor de entrada utilizado por la función.

    Devuelve
    --------
    Any
        Resultado generado por la función.
    """
    return [
        str(celda.value).strip() if celda.value is not None else ""
        for celda in hoja[3]
    ]


def _clave_fila(valores):
    """
    Ejecuta la lógica asociada a clave fila.

    Parámetros
    ----------
    valores : Any
        Valores que se van a procesar.

    Devuelve
    --------
    Any
        Resultado generado por la función.
    """
    clave = []
    for valor in valores:
        if isinstance(valor, datetime):
            clave.append(valor.replace(microsecond=0).isoformat())
        else:
            clave.append(str(valor) if valor is not None else "")
    return tuple(clave)


def _recoger_filas(rutas, nombre_hoja, cabeceras_destino, inicio, fin, paciente_id, sesion_id):
    """
    Ejecuta la lógica asociada a recoger filas.

    Parámetros
    ----------
    rutas : Any
        Conjunto de rutas que se van a procesar.

    nombre_hoja : Any
        Valor de entrada utilizado por la función.

    cabeceras_destino : Any
        Valor de entrada utilizado por la función.

    inicio : Any
        Instante inicial del intervalo.

    fin : Any
        Instante final del intervalo.

    paciente_id : Any
        Identificador del paciente.

    sesion_id : Any
        Identificador de la sesión.

    Devuelve
    --------
    Any
        Resultado generado por la función.
    """
    filas = []
    vistos = set()
    for ruta in rutas:
        with abrir_libro(ruta, read_only=True, data_only=False) as libro:
            hoja = _buscar_hoja(libro, nombre_hoja)
            if hoja is None:
                continue
            cabeceras_origen = _cabeceras(hoja)
            indices = {
                nombre: indice
                for indice, nombre in enumerate(cabeceras_origen)
                if nombre
            }
            if "timestamp" not in indices:
                continue

            for valores_origen in hoja.iter_rows(min_row=4, values_only=True):
                instante = _parsear_fecha(valores_origen[indices["timestamp"]])
                if instante is None or instante < inicio or instante > fin:
                    continue

                valores = []
                for nombre in cabeceras_destino:
                    if nombre == "paciente_id":
                        valor = paciente_id
                    elif nombre == "sesion_bis_id":
                        valor = sesion_id
                    elif nombre in indices and indices[nombre] < len(valores_origen):
                        valor = valores_origen[indices[nombre]]
                    else:
                        valor = None
                    valores.append(valor)

                clave = _clave_fila(valores)
                if clave not in vistos:
                    vistos.add(clave)
                    filas.append(valores)

    indice_timestamp = cabeceras_destino.index("timestamp")
    filas.sort(key=lambda fila: _parsear_fecha(fila[indice_timestamp]) or datetime.max)
    return filas


def _capturar_estilo_fila(hoja, fila, numero_columnas):
    """
    Ejecuta la lógica asociada a capturar estilo fila.

    Parámetros
    ----------
    hoja : Any
        Valor de entrada utilizado por la función.

    fila : Any
        Valor de entrada utilizado por la función.

    numero_columnas : Any
        Valor de entrada utilizado por la función.

    Devuelve
    --------
    Any
        Resultado generado por la función.
    """
    return [
        {
            "style": copy(hoja.cell(fila, columna)._style),
            "number_format": hoja.cell(fila, columna).number_format,
            "alignment": copy(hoja.cell(fila, columna).alignment),
        }
        for columna in range(1, numero_columnas + 1)
    ]


def _reemplazar_datos(hoja, filas):
    """
    Ejecuta la lógica asociada a reemplazar datos.

    Parámetros
    ----------
    hoja : Any
        Valor de entrada utilizado por la función.

    filas : Any
        Valor de entrada utilizado por la función.

    Devuelve
    --------
    None
        La función no devuelve ningún valor.
    """
    cabeceras = _cabeceras(hoja)
    numero_columnas = len(cabeceras)
    estilo = _capturar_estilo_fila(hoja, 4, numero_columnas)

    if hoja.max_row >= 4:
        hoja.delete_rows(4, hoja.max_row - 3)

    filas_escritas = filas if filas else [[None] * numero_columnas]
    for indice_fila, valores in enumerate(filas_escritas, start=4):
        for indice_columna, valor in enumerate(valores, start=1):
            celda = hoja.cell(indice_fila, indice_columna, valor)
            plantilla = estilo[indice_columna - 1]
            celda._style = copy(plantilla["style"])
            celda.number_format = plantilla["number_format"]
            celda.alignment = copy(plantilla["alignment"])

    ultima_fila = 3 + len(filas_escritas)
    ultima_columna = get_column_letter(numero_columnas)
    for tabla in hoja.tables.values():
        tabla.ref = f"A3:{ultima_columna}{ultima_fila}"


def _actualizar_general(libro, paciente_id, carpeta_paciente, sesion_id):
    """
    Actualiza general.

    Parámetros
    ----------
    libro : Any
        Valor de entrada utilizado por la función.

    paciente_id : Any
        Identificador del paciente.

    carpeta_paciente : Any
        Valor de entrada utilizado por la función.

    sesion_id : Any
        Identificador de la sesión.

    Devuelve
    --------
    None
        La función no devuelve ningún valor.
    """
    hoja = _buscar_hoja(libro, "general")
    if hoja is None:
        return
    cabeceras = {
        str(celda.value).strip(): celda.column
        for celda in hoja[3]
        if celda.value is not None
    }
    valores = {
        "paciente_id": paciente_id,
        "carpeta_paciente": carpeta_paciente,
        "sesion_bis_id": sesion_id,
    }
    for nombre, valor in valores.items():
        if nombre in cabeceras:
            hoja.cell(4, cabeceras[nombre], valor)


def _crear_metadata(libro, paciente_id, sesion, rutas_icca):
    """
    Crea metadata.

    Parámetros
    ----------
    libro : Any
        Valor de entrada utilizado por la función.

    paciente_id : Any
        Identificador del paciente.

    sesion : Any
        Valor de entrada utilizado por la función.

    rutas_icca : Any
        Valor de entrada utilizado por la función.

    Devuelve
    --------
    None
        La función no devuelve ningún valor.
    """
    if "metadata_sesion" in libro.sheetnames:
        del libro["metadata_sesion"]
    hoja = libro.create_sheet("metadata_sesion", 1)
    hoja.append(["Campo", "Valor"])
    hoja.append(["Paciente", paciente_id])
    hoja.append(["Sesion BIS", sesion["sesion_id"]])
    hoja.append(["Inicio BIS", _parsear_fecha(sesion["inicio"])])
    hoja.append(["Fin BIS", _parsear_fecha(sesion["fin"])])
    hoja.append(["Modo BIS", sesion.get("modo", "")])
    hoja.append(["Excel ICCA origen", "; ".join(str(Path(ruta).name) for ruta in rutas_icca)])
    hoja.append(["Generado", datetime.now().replace(microsecond=0)])

    hoja["A1"].font = Font(bold=True, color="FFFFFF")
    hoja["B1"].font = Font(bold=True, color="FFFFFF")
    hoja["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    hoja["B1"].fill = PatternFill("solid", fgColor="1F4E78")
    hoja.column_dimensions["A"].width = 24
    hoja.column_dimensions["B"].width = 70
    hoja.freeze_panes = "A2"
    for fila in (4, 5, 8):
        hoja.cell(fila, 2).number_format = "dd/mm/yyyy hh:mm:ss"


def generar_excel_icca_sesion(
    rutas_icca,
    sesion,
    salida,
    paciente_id,
    carpeta_paciente,
):
    """
    Genera excel icca sesion.

    Parámetros
    ----------
    rutas_icca : Any
        Valor de entrada utilizado por la función.

    sesion : Any
        Valor de entrada utilizado por la función.

    salida : Any
        Valor de entrada utilizado por la función.

    paciente_id : Any
        Identificador del paciente.

    carpeta_paciente : Any
        Valor de entrada utilizado por la función.

    Devuelve
    --------
    Any
        Resultado generado por la función.

    Lanza
    -----
    ValueError
        Si se produce una condición no válida durante la ejecución.
    """
    if not rutas_icca:
        raise ValueError("No hay archivos ICCA compatibles con la sesion BIS.")

    inicio = _parsear_fecha(sesion["inicio"])
    fin = _parsear_fecha(sesion["fin"])
    salida = Path(salida)
    salida.parent.mkdir(parents=True, exist_ok=True)

    with abrir_libro(rutas_icca[0]) as libro:
        _actualizar_general(
            libro,
            paciente_id=paciente_id,
            carpeta_paciente=carpeta_paciente,
            sesion_id=sesion["sesion_id"],
        )
        for nombre in HOJAS_TEMPORALES:
            hoja = _buscar_hoja(libro, nombre)
            if hoja is None:
                continue
            cabeceras = _cabeceras(hoja)
            filas = _recoger_filas(
                rutas_icca,
                nombre,
                cabeceras,
                inicio,
                fin,
                paciente_id,
                sesion["sesion_id"],
            )
            _reemplazar_datos(hoja, filas)

        _crear_metadata(libro, paciente_id, sesion, rutas_icca)
        libro.save(salida)
    return salida
